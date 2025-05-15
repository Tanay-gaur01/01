import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Initialize session state for feedback
if 'feedback' not in st.session_state:
    st.session_state.feedback = []

def add_feedback(feedback_text):
    st.session_state.feedback.append(feedback_text)

def download_feedback():
    df = pd.DataFrame(st.session_state.feedback, columns=['Feedback'])
    df.to_csv('feedback.csv', index=False)
    st.success("Feedback downloaded successfully!")

def remove_duplicates(source_df):
    duplicates = source_df.duplicated(subset=['Revised Copy'], keep='first')
    if duplicates.any():
        st.info("Duplicates found in Source Sheet. Removing duplicates...")
        source_df = source_df.drop_duplicates(subset=['Revised Copy'], keep='first')
    else:
        st.info("No duplicates found in 'Revised Copy'.")
    return source_df

def map_content(source_df, site_df):
    source_df = remove_duplicates(source_df)
    source_dict = {row['Design Copy']: (index, row['Revised Copy']) for index, row in source_df.iterrows()}

    # Add new columns if they don't exist
    for col in ['Mapped Cell', 'Revised Copy']:
        if col not in site_df.columns:
            site_df[col] = ''

    for index, row in site_df.iterrows():
        design_copy = row['Design Copy']
        if design_copy in source_dict:
            source_index, revised_copy = source_dict[design_copy]
            site_df.at[index, 'Mapped Cell'] = f"Source!A{source_index + 2}"
            site_df.at[index, 'Revised Copy'] = str(revised_copy)
        else:
            site_df.at[index, 'Mapped Cell'] = 'Not Found'

    return site_df

def structure_and_format_data(raw_data, group_column='frame'):
    # Get unique frames in original order
    frame_order = raw_data[group_column].dropna().drop_duplicates().tolist()

    # Create structured DataFrame
    structured_data = pd.DataFrame(columns=raw_data.columns)

    for name in frame_order:
        # Filter group while preserving original order
        group = raw_data[raw_data[group_column] == name].reset_index(drop=True)

        # Add title row
        title_row = pd.DataFrame([[name] + [''] * (len(raw_data.columns) - 1)],
                                 columns=raw_data.columns)
        structured_data = pd.concat([structured_data, title_row], ignore_index=True)

        # Add empty row
        structured_data = pd.concat([structured_data,
                                   pd.DataFrame([[''] * len(raw_data.columns)],
                                   columns=raw_data.columns)], ignore_index=True)

        # Add group data
        structured_data = pd.concat([structured_data, group], ignore_index=True)

        # Add empty row
        structured_data = pd.concat([structured_data,
                                   pd.DataFrame([[''] * len(raw_data.columns)],
                                   columns=raw_data.columns)], ignore_index=True)

    # Remove last empty row
    return structured_data.iloc[:-1]

def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Workbook styling
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Apply yellow background to title rows
    for idx, row in enumerate(worksheet.iter_rows(min_row=1), 1):
        if worksheet.cell(row=idx, column=1).value in df['frame'].dropna().unique():
            for cell in row:
                cell.fill = fill

    writer.close()
    return output.getvalue()

# Streamlit UI
st.title("Excel Content Mapper")

st.write("""
### Upload your files
1. **Source File** - Contains original and revised copies
2. **Site File** - File to be updated with mapping information
""")

source_file = st.file_uploader("Upload Source Excel File", type=['xlsx'])
site_file = st.file_uploader("Upload Site Excel File", type=['xlsx'])

if source_file and site_file:
    source_df = pd.read_excel(source_file)
    site_df = pd.read_excel(site_file, dtype=str)

    # Mapping step
    if st.button("Map Data"):
        try:
            with st.spinner('Mapping content...'):
                mapped_df = map_content(source_df, site_df)
                st.session_state.mapped_df = mapped_df

            st.success("Mapping completed!")
            st.write("### Mapped Data Preview")
            st.dataframe(mapped_df.head(10))

        except KeyError as e:
            st.error(f"Missing required column: {e}")
        except Exception as e:
            st.error(f"Error during mapping: {str(e)}")

    # Structuring step
    if st.button("Structure Data") and 'mapped_df' in st.session_state:
        try:
            with st.spinner('Structuring data...'):
                structured_df = structure_and_format_data(st.session_state.mapped_df)
                st.session_state.structured_df = structured_df

            st.success("Structuring completed!")
            st.write("### Structured Data Preview")
            st.dataframe(structured_df.head(15))

        except KeyError as e:
            st.error(f"Missing required column: {e}")
        except Exception as e:
            st.error(f"Error during structuring: {str(e)}")

    # Download step
    if 'structured_df' in st.session_state:
        st.write("### Download Final File")
        output = to_excel(st.session_state.structured_df)

        original_name = site_file.name.split('.')[0]
        filename = f"{original_name}_Mapped_Structured.xlsx"

        st.download_button(
            label="Download Processed File",
            data=output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# Feedback Section
st.title("User Feedback")

# Text area for entering feedback
feedback_text = st.text_area("Enter your feedback or review:")

# Button to add feedback
if st.button("Submit Feedback"):
    if feedback_text:
        add_feedback(feedback_text)
        st.success("Thank you for your feedback!")
    else:
        st.error("Please enter your feedback.")

# Admin section to download feedback
st.title("Admin Panel")

# Button to download feedback
if st.button("Download Feedback"):
    if st.session_state.feedback:
        download_feedback()
    else:
        st.error("No feedback to download.")

# Display current feedback
st.subheader("Current Feedback")
for idx, feedback in enumerate(st.session_state.feedback, start=1):
    st.write(f"{idx}. {feedback}")
