import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def remove_duplicates(source_df):
    duplicates = source_df.duplicated(subset=['Revised Copy'], keep='first')
    if duplicates.any():
        st.info("Duplicates found in Source Sheet. Removing duplicates...")
        source_df = source_df.drop_duplicates(subset=['Revised Copy'], keep='first')
    else:
        st.info("No duplicates found in 'Revised Copy'.")
    return source_df

def map_content(source_df, site_df):
    source_df = remove_duplicates(source_df)
    source_dict = {row['Design Copy']: (index, row['Revised Copy']) for index, row in source_df.iterrows()}

    for col in ['Mapped Cell', 'Revised Copy']:
        if col not in site_df.columns:
            site_df[col] = ''

    for index, row in site_df.iterrows():
        design_copy = row['Design Copy']
        if design_copy in source_dict:
            source_index, revised_copy = source_dict[design_copy]
            site_df.at[index, 'Mapped Cell'] = f"Source!A{source_index + 2}"
            site_df.at[index, 'Revised Copy'] = str(revised_copy)
        else:
            site_df.at[index, 'Mapped Cell'] = 'Not Found'

    return site_df

def structure_and_format_data(raw_data, group_column='frame'):
    frame_order = raw_data[group_column].dropna().drop_duplicates().tolist()
    structured_data = pd.DataFrame(columns=raw_data.columns)

    for name in frame_order:
        group = raw_data[raw_data[group_column] == name].reset_index(drop=True)
        
        title_row = pd.DataFrame([[name] + [''] * (len(raw_data.columns) - 1)],
                               columns=raw_data.columns)
        structured_data = pd.concat([structured_data, title_row], ignore_index=True)
        structured_data = pd.concat([structured_data,
                                   pd.DataFrame([[''] * len(raw_data.columns)],
                                   columns=raw_data.columns)], ignore_index=True)
        structured_data = pd.concat([structured_data, group], ignore_index=True)
        structured_data = pd.concat([structured_data,
                                   pd.DataFrame([[''] * len(raw_data.columns)],
                                   columns=raw_data.columns)], ignore_index=True)

    return structured_data.iloc[:-1]

def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for idx, row in enumerate(worksheet.iter_rows(min_row=1), 1):
        if worksheet.cell(row=idx, column=1).value in df['frame'].dropna().unique():
            for cell in row:
                cell.fill = fill

    writer.close()
    return output.getvalue()
