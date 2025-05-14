import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def remove_duplicates(source_df):
    duplicates = source_df.duplicated(subset=['Revised Copy'], keep='first')
    if duplicates.any():
        st.info("Duplicates found in Source Sheet. Removing duplicates...")
        source_df = source_df.drop_duplicates(subset=['Revised Copy'], keep='first')
    else:
        st.info("No duplicates found in 'Revised Copy'.")
    return source_df

def map_content(source_df, site_df):
    source_df = remove_duplicates(source_df)
    source_dict = {row['Design Copy']: (index, row['Revised Copy']) for index, row in source_df.iterrows()}

    # Add 'Mapped Cell' column if it doesn't exist
    if 'Mapped Cell' not in site_df.columns:
        site_df['Mapped Cell'] = ''

    # Add 'Revised Copy' column if it doesn't exist
    if 'Revised Copy' not in site_df.columns:
        site_df['Revised Copy'] = ''
    
    for index, row in site_df.iterrows():
        design_copy = row['Design Copy']
        if design_copy in source_dict:
            source_index, revised_copy = source_dict[design_copy]
            site_df.at[index, 'Mapped Cell'] = f"Source!A{source_index + 2}"
            site_df.at[index, 'Revised Copy'] = str(revised_copy)
        else:
            site_df.at[index, 'Mapped Cell'] = 'Not Found'

    return site_df

def structure_and_format_data(raw_data, group_column='frame'):
    # Get unique frames in their original order of appearance
    frame_order = raw_data[group_column].dropna().drop_duplicates().tolist()

    # Group data by the specified column WITHOUT sorting
    grouped_data = raw_data.groupby(group_column, as_index=False, sort=False)

    # Create structured DataFrame
    structured_data = pd.DataFrame(columns=raw_data.columns)

    # Build structured data with title rows and empty rows
    for name, group in grouped_data:
        # Add frame title row
        title_row = pd.DataFrame([[name] + [''] * (len(raw_data.columns) - 1)],
                                 columns=raw_data.columns)
        structured_data = pd.concat([structured_data, title_row], ignore_index=True)

        # Add empty row
        structured_data = pd.concat([structured_data,
                                     pd.DataFrame([[''] * len(raw_data.columns)],
                                     columns=raw_data.columns)], ignore_index=True)

        # Add group data
        structured_data = pd.concat([structured_data, group], ignore_index=True)

        # Add empty row
        structured_data = pd.concat([structured_data,
                                     pd.DataFrame([[''] * len(raw_data.columns)],
                                     columns=raw_data.columns)], ignore_index=True)

    # Remove last empty row
    structured_data = structured_data.iloc[:-1]

    return structured_data

def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Access the workbook and the worksheet for styling
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Define a fill style for the title rows
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Apply the fill to the title rows
    for row in worksheet.iter_rows():
        if row[0].value in df['frame'].dropna().unique():
            for cell in row:
                cell.fill = fill

    writer.close()
    processed_data = output.getvalue()
    return processed_data

# Streamlit UI
st.title("Excel Content Mapper")

st.write("""
### Upload your files
1. **Source File** - Contains the original and revised copies
2. **Site File** - File to be updated with mapping information
""")

source_file = st.file_uploader("Upload Source Excel File", type=['xlsx'])
site_file = st.file_uploader("Upload Site Excel File", type=['xlsx'])

if source_file and site_file:
    # Read files
    source_df = pd.read_excel(source_file)
    site_df = pd.read_excel(site_file)

    # Button to structure the data
    if st.button("Structure Data"):
        try:
            with st.spinner('Structuring files...'):
                structured_site_df = structure_and_format_data(site_df, group_column='frame')

            st.success("File structuring completed!")

            st.write("### Preview of Structured Data")
            st.dataframe(structured_site_df.head())

            # Store structured data in session state
            st.session_state['structured_site_df'] = structured_site_df

        except KeyError as e:
            st.error(f"Missing required column: {e}")
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")

    # Button to start mapping
    if st.button("Start Mapping") and 'structured_site_df' in st.session_state:
        try:
            # Process files
            with st.spinner('Mapping files...'):
                updated_site_df = map_content(source_df, st.session_state['structured_site_df'])

            # Download button
            st.success("File mapping completed!")

            st.write("### Preview of Mapped Data")
            st.dataframe(updated_site_df.head())

            # Generate the output filename
            original_filename = site_file.name.replace('.xlsx', '')
            output_filename = f"Mapped_{original_filename}.xlsx"

            st.write("### Download Final File")
            excel_data = to_excel(updated_site_df)
            st.download_button(
                label="Download Final File",
                data=excel_data,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except KeyError as e:
            st.error(f"Missing required column: {e}")
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")